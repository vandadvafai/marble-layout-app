"""Portability regressions for the V1.1 clean-install pass.

These tests simulate what happens on a fresh laptop where none of
the generated files under ``outputs/`` exist and no env override is
set. The app must:

  * boot and let the user pick a sample plan (bundled fixtures),
  * return the ``empty`` inventory label with zero counts (not 500),
  * refuse real-project exports with a clean 400 (not a crash),
  * accept an upload and swap in a working inventory,
  * keep two projects isolated so the second upload doesn't inherit
    slab state from the first.

Every test resets ``AVANDAD_INVENTORY_PATH`` /
``STONELAYOUT_INVENTORY_PATH`` and calls ``clear_active_upload()``
so it stands alone regardless of what conftest.py did.
"""
from __future__ import annotations

import io
from pathlib import Path

import pytest
from fastapi.testclient import TestClient

from placement_engine.api.main import app
from placement_engine.api.inventory_upload import clear_active_upload


client = TestClient(app)


@pytest.fixture(autouse=True)
def _reset_state(monkeypatch):
    """Every test in this module starts with no env override and no
    uploaded session — the "fresh laptop" baseline."""
    monkeypatch.delenv("AVANDAD_INVENTORY_PATH", raising=False)
    monkeypatch.delenv("STONELAYOUT_INVENTORY_PATH", raising=False)
    clear_active_upload()
    yield
    clear_active_upload()


# ---------------------------------------------------------------------------
# Empty inventory portability
# ---------------------------------------------------------------------------


def test_inventory_info_empty_on_fresh_install():
    """A brand-new clone with no upload and no env override must
    return the ``empty`` state — NOT a 500 pointing at a
    developer-only ``outputs/`` path."""
    r = client.get("/api/inventory/info")
    assert r.status_code == 200, r.text
    body = r.json()
    assert body["source_label"] == "empty"
    assert body["source_path"] is None
    assert body["valid_count"] == 0
    assert body["total_records"] == 0
    assert body["stats"] is None


def test_match_inventory_empty_state_returns_no_match_per_piece():
    """The matcher must not crash when no inventory is uploaded. It
    reports ``no_match`` for every piece with an empty candidate
    list so the frontend can render its blocker instead of a 500."""
    r = client.post(
        "/api/demo-layouts/l_shape/match-inventory",
        json={
            "pieces": [
                {"piece_id": "p1", "nominal_width_mm": 500,
                 "nominal_height_mm": 500},
                {"piece_id": "p2", "nominal_width_mm": 700,
                 "nominal_height_mm": 900},
            ],
        },
    )
    assert r.status_code == 200, r.text
    body = r.json()
    assert body["inventory"]["source_label"] == "empty"
    assert body["inventory_count"] == 0
    assert body["summary"]["no_match"] == 2
    for pe in body["pieces"]:
        assert pe["status"] == "no_match"
        assert pe["candidates"] == []


def test_export_dxf_refuses_without_uploaded_inventory():
    """The factory-DXF endpoint requires slab metadata. On a fresh
    install it must fail with a clear 400 telling the operator to
    finish Step 3 first — not crash inside the loader."""
    r = client.post(
        "/api/demo-layouts/l_shape/export-dxf",
        json={
            "pieces": [{
                "piece_id": "p1",
                "polygon": [[0, 0], [500, 0], [500, 500], [0, 500], [0, 0]],
                "nominal_width_mm": 500.0,
                "nominal_height_mm": 500.0,
            }],
            "assignments": {"p1": "some-slab"},
        },
    )
    assert r.status_code == 400, r.text
    detail = r.json()["detail"]
    assert "No inventory uploaded" in (
        detail if isinstance(detail, str) else detail.get("message", "")
    ), detail


def test_regenerate_layout_uses_bundled_sample_for_sample_plans():
    """Sample plans keep the ``inventory median`` regeneration
    working via the bundled ``examples/demo/clean_slabs.json`` even
    when the operator hasn't uploaded anything. Real projects gate
    Step 4 at the frontend so this never leaks into a shipped DXF
    (see the frontend workflow gate + the export refuses tests)."""
    r = client.post(
        "/api/demo-layouts/l_shape/regenerate",
        json={},
    )
    assert r.status_code == 200, r.text
    body = r.json()
    assert body["tile_choice"]["basis"] == "inventory_median"
    assert body["inventory_source_label"] == "empty"


def test_regenerate_layout_accepts_explicit_tile_without_inventory():
    """Even without an upload, the caller can regenerate the layout
    when they supply explicit tile dimensions — the sample-plan
    workflow can then proceed without demanding Step 3."""
    r = client.post(
        "/api/demo-layouts/l_shape/regenerate",
        json={"tile_width_mm": 1500.0, "tile_height_mm": 2500.0},
    )
    assert r.status_code == 200, r.text
    body = r.json()
    assert body["tile_choice"]["basis"] == "explicit_override"
    assert body["tile_choice"]["tile_width_mm"] == 1500.0


# ---------------------------------------------------------------------------
# Sample plans still work without any inventory
# ---------------------------------------------------------------------------


def test_sample_plans_still_load_without_uploaded_inventory():
    """``GET /api/demo-layouts/{id}`` seeds its tile size from the
    BUNDLED sample inventory (``examples/demo/clean_slabs.json``),
    NEVER from ``outputs/``. It must succeed on a fresh clone."""
    for demo_id in ("l_shape", "rectangle", "apartment"):
        r = client.get(f"/api/demo-layouts/{demo_id}")
        assert r.status_code == 200, (demo_id, r.text)
        body = r.json()
        assert body["layout"]["piece_count"] > 0


# ---------------------------------------------------------------------------
# Excel column validation
# ---------------------------------------------------------------------------


def _build_english_excel_bytes() -> tuple[bytes, str]:
    """Compose a tiny .xlsx in memory using openpyxl with headers a
    non-Persian ERP might export. Two slabs, valid dimensions."""
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.append(["Serial Number", "Width (cm)", "Height (cm)", "Item Code"])
    ws.append(["ABC-01", 150, 200, "P-1"])
    ws.append(["ABC-02", 160, 220, "P-2"])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue(), "english_export.xlsx"


def _build_missing_columns_excel_bytes() -> tuple[bytes, str]:
    """Excel with no identity column at all — just an area figure.
    Must be rejected at upload time with a listing of the missing
    columns (identity + dimensions)."""
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.append(["Area", "Colour"])
    ws.append([3.0, "beige"])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue(), "no_id.xlsx"


def test_upload_accepts_english_column_names():
    """An English-header Excel matches Persian aliases in the column
    map — a non-Persian ERP can drive the upload endpoint without
    the operator translating headers first."""
    data, name = _build_english_excel_bytes()
    r = client.post(
        "/api/inventory/upload",
        files={
            "excel": (name, data,
                      "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
        },
    )
    assert r.status_code == 200, r.text
    body = r.json()
    assert body["summary"]["valid_slabs"] == 2


def test_upload_rejects_missing_required_columns():
    """The upload endpoint surfaces which columns weren't recognised
    so the operator can fix the Excel instead of guessing."""
    data, name = _build_missing_columns_excel_bytes()
    r = client.post(
        "/api/inventory/upload",
        files={
            "excel": (name, data,
                      "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
        },
    )
    assert r.status_code == 400
    detail = r.json()["detail"]
    assert detail["error"] == "invalid_excel"
    msg = detail["message"]
    assert "missing required columns" in msg.lower()
    assert "identity" in msg.lower()


# ---------------------------------------------------------------------------
# Two independent projects — state isolation
# ---------------------------------------------------------------------------


def test_two_projects_do_not_share_slabs():
    """Upload project A, then upload project B, then confirm project
    A's slabs are no longer visible. ``clear_active_upload()`` is
    called between projects the same way ``Start new project`` does
    in the frontend."""
    from openpyxl import Workbook

    def _make(rows):
        wb = Workbook()
        ws = wb.active
        ws.append(["Serial", "Width (cm)", "Height (cm)"])
        for r in rows:
            ws.append(r)
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    project_a = _make([["A-1", 150, 200]])
    project_b = _make([["B-1", 160, 220], ["B-2", 170, 230]])

    r_a = client.post(
        "/api/inventory/upload",
        files={"excel": ("a.xlsx", project_a,
                         "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert r_a.status_code == 200
    info_a = client.get("/api/inventory/info").json()
    assert info_a["valid_count"] == 1

    # Start new project — the frontend hits this endpoint too.
    r_del = client.delete("/api/inventory/current")
    assert r_del.status_code == 200

    r_b = client.post(
        "/api/inventory/upload",
        files={"excel": ("b.xlsx", project_b,
                         "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert r_b.status_code == 200
    info_b = client.get("/api/inventory/info").json()
    assert info_b["valid_count"] == 2
    # And project A's slab id must NOT be reachable — the matcher
    # only sees project B's inventory.
    ids = {
        r["slab_id"]
        for r in client.get("/api/inventory/current").json()
                     .get("summary", {})
                     .get("preview", [])
    }
    assert not (ids & {"A-1"})


# ---------------------------------------------------------------------------
# APP_DATA_DIR resolution
# ---------------------------------------------------------------------------


def test_app_paths_created_on_demand(tmp_path, monkeypatch):
    """``ensure_dirs`` must create every configured subdirectory the
    first time it's called — no manual bootstrap required."""
    from placement_engine.api.app_paths import (
        AppPaths, ensure_dirs, resolve_app_paths,
    )
    monkeypatch.setenv("AVANDAD_DATA_DIR", str(tmp_path / "data"))
    paths: AppPaths = resolve_app_paths()
    # Nothing exists yet.
    assert not paths.uploads.exists()
    ensure_dirs(paths)
    assert paths.uploads.exists()
    assert paths.processed.exists()
    assert paths.exports.exists()
